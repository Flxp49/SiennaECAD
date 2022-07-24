
from cmath import log
from openpyxl import load_workbook
import xlrd
import pandas as pd
files = []
dfs = []

# =========================================================================================


def createDataFrame(data):
    dfs.append(pd.DataFrame(data, columns=["Net Name", "Net Node"]))
    print(dfs[0])

# ==============================================================================================

# For altium.net


def processNet(filename):
    original_data = list()
    final_data = list()
    c = 0
    with open(filename, "r") as file:
        for line in file.readlines():
            if line.strip() != "":
                original_data.append(line.strip())

    for i in range(len(original_data)):
        if original_data[i] == "(":
            original_data = original_data[i:]
            break

    for i in range(0, len(original_data)):
        if original_data[i] not in ["(", ")"] and c == 0:
            column_one = original_data[i].replace('"', '')
            c = c+1
        elif original_data[i] not in ["(", ")"] and c != 0:
            column_two = original_data[i].replace("-", ".").replace('"', '')
            final_data.append([column_one, column_two])
        else:
            c = 0

    createDataFrame(final_data)
# processNet("Netlist/Altium/Altium.net")

# for SCH.dat
def processDat(filename):
    original_data = list()
    final_data = list()
    with open(filename, "r") as file:
        for line in file.readlines():
            if line.strip() != "":
                 original_data.append(line.strip())
        i = 0
        while (i < len(original_data)):
            if original_data[i] == "NET_NAME":
                 column_one = original_data[i+1].replace("'", '')
                 i=i+2
                 continue
            if "NODE_NAME" in original_data[i]:
                column_two = original_data[i].split()[1] + "/" + original_data[i].split()[2]
                final_data.append([column_one, column_two])
                i=i+2
                continue
            i=i+1
        print(final_data)

# processDat("Netlist/Allegro/SCH.dat")


# for SCH.asc or PCB.asc
def processAsc(filename):

    def findAscType(data):
        for i in range(len(data)):
            if data[i] == "*ROUTE*":
                return 1, i # this is _PCB.asc
            if data[i] == "*NET*":
                return 2, i # this is _SCH.asc
        return -1
    
    original_data = list()
    final_data = list()
    with open(filename, "r") as file:
        for line in file.readlines():
            if line.strip() != "":
                original_data.append(line.strip())
    ascType = findAscType(original_data)
    if ascType == -1:
        print("File type not found")
    else:
        original_data = original_data[ascType[1]:]
        if (ascType[0] == 1):
            print(original_data[:20])
            # if line.strip() != "" and (" 1356" not in line.strip() or "THERMAL" not in line.strip()):
            #      original_data.append(line.strip())

processAsc("Netlist/Pads/PCB.asc")
# processAsc("Netlist/Pads/SCH.asc")



def processText(filename):
    original_data = list()
    with open(filename, "r") as File:
        for line in File.readlines():
            temp_list = line.split()
            if temp_list != []:
                original_data.append(line.split())

    # Find out what type text file it is
    textTypeloc = getTextType(original_data)
    if textTypeloc[0] == 1:
        cdata = original_data[textTypeloc[1]:-1]
        fdata = getMappings(cdata, textTypeloc[0])
    elif textTypeloc[0] == 2:
        cdata = original_data[textTypeloc[1]:-1]
        fdata = getMappings(cdata, textTypeloc[0])
    else:
        print("ERROR, TextType unknown.")
        return


def getTextType(data):
    for i in range(len(data)):
        if ".ADD_TER" in data[i]:
            return 1, i
        elif "*" in data[i][0]:
            return 2, i
    return -1, -1


def getMappings(data, type):
    final_data = list()
    column_one = ""
    column_two = ""

    if type == 1:
        for i in data:
            # Updating column one if the row is .ADD_TER
            if ".ADD_TER" in i:
                column_one = i[3]
                column_two = i[1] + i[2]

            # Updating the column two details
            elif ".TER" in i:
                column_two = i[1] + i[2]
            else:
                column_two = i[0] + i[1]

            temp_dict = {
                "Net Name": column_one.replace('"', ''),
                "Component Name": (column_two.replace("-", ".")).replace('"', '')
            }
            final_data.append(temp_dict)
        return final_data

    elif type == 2:
        for i in data:
            # Updating column one if the row is *
            if "*" in i[0]:
                column_one = i[0] + " " + i[1]
            else:
                for j in i:
                    column_two = j
                    temp_dict = {
                        "Net Name": column_one.replace('"', ''),
                        "Component Name": (column_two.replace("-", ".")).replace('"', '')
                    }
                    final_data.append(temp_dict)
        return final_data


# n = int(input("Enter the no files to convert: "))
# for i in range(n):
#     files.append(input("Enter filename" + str(i+1)+": "))
# for file in files:
#     fileformat = file.split(".")[1]
#     if fileformat == "txt":
#         processText(file)
#     elif fileformat == "xls" or fileformat == "xlsx":
#         processXls(file, fileformat)
#     elif fileformat == "net":
#         processNet(file)

# for AllegroPCB.xls
# def processXls(filename):
    # pd.read_excel(filename, usecols="A,B", skiprows=3, engine="openpyxl")
    # print(pd)
    # final_data = list()
    # wb = xlrd.open_workbook(filename)
    # sheet = wb.sheet_by_index(0)
    # for i in range(5, sheet.nrows):
    #     print(sheet.cell_value(i, 0))
    # column_one = sheet.cell_value(i, 0).replace('"', '')
    # cmps = sheet.cell_value(i, 1).split()
    # for j in cmps:
    #     column_two = j.replace("-", ".").replace('"', '')
    #     final_data.append([column_one, column_two])

    # print(final_data)
    # createDataFrame(final_data)

    # processNet("Netlist/Allegro/PCB.xlsx")

    # elif type == 2:
    #     wb = xlrd.open_workbook(filename)
    #     sheet = wb.sheet_by_index(0)
    #     for i in range(3, sheet.nrows):
    #         if (sheet.cell_value(i, 0) != ""):
    #             column_one = sheet.cell_value(i, 0)
    #             continue
    #         elif (sheet.cell_value(i, 0) == "" and sheet.cell_value(i, 3) == ""):
    #             continue
    #         elif (sheet.cell_value(i, 0) == ""):
    #             column_two = sheet.cell_value(i, 3)
    #             temp_dict = {
    #                 "Net Name": column_one.replace('"', ''),
    #                 "Component Name": (column_two.replace("-", ".")).replace('"', '')
    #             }
    #             final_data.append(temp_dict)
    # else:
    #     print("Invalid input, Please enter 1 or 2 for xls files")
    #     return

    # else:
    #     if type == 2 or 3:
    #         wb = load_workbook(filename)
    #         sheet = wb[wb.sheetnames[0]]
    #         s = 2 if type == 2 else 6
    #         for row in range(s, sheet.max_row+1):
    #             for column in "AB":
    #                 cell = "{}{}".format(column, row)
    #                 if column == "A" and sheet[cell].value != None:
    #                     column_one = sheet[cell].value
    #                 elif column == "B" and sheet[cell].value != None:
    #                     cmps = sheet[cell].value.split()
    #                     for j in cmps:
    #                         column_two = j
    #                         temp_dict = {
    #                             "Net Name": column_one.replace('"', ''),
    #                             "Component Name": (column_two.replace("-", ".")).replace('"', '')
    #                         }
    #                         final_data.append(temp_dict)

    # else:
    #     print("Invalid input, Please enter 2 or 3 for xlsx files")
    #     return
